#encoding=utf-8
#author: blue16（Index:blue16.cn）
#date: 2023-10-18
#summary: 这个才是正式的3.0版本，修复了3.0版本周数识别错误的问题，可以手动选择生成无课表还是有课表
#遍历学生会所有部门和对应的的课程表（列表样式）并自动生成无课表(加入了周期范围分析)


import tabula
import csv
import openpyxl
import os
import re

number_detail=0
number_weekrange=0
list_failed=[]
list_courserange=[]
list_weekrange=[]

#将一个pdf转为csv
def convert_to_csv(pdf_path, csv_path):
    tabula.convert_into(pdf_path, csv_path, output_format="csv", pages="all")

#给定一个csv，遍历搜索所有课程明细，返回一个list
def get_detail(csv_path):
    with open(csv_path,encoding='utf-8') as csvfile:
        content=csvfile.read()
        #print(content)
        #print(content.split(","))
        res=[]
        for temp in content.split(","):
            if("周数" in temp):
                res.append(temp)
        print("课程明细总数",len(res))
        global number_detail
        number_detail=len(res)
        return res

#对于单个课程明细（一个String），返回正则表达式匹配到的文本块，类似于“周数：A-B”，也就是提取课程周次，EG：['9-10', '11']
def get_weekrange(str):
    section=re.findall("(?<=周数: ).*?(?=周)",str)
    return section

#读取指定的列,并清除无效数据（这里是用于读取第二列，也就是节次范围那一列）
def get_column(csv_path,n):#csv_path=example.csv
    with open(csv_path,encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        templist=[]
        for row in reader:
            temp=row[n]
            #print(temp,len(temp))
            if len(temp)>=3 and len(temp)<=5:#加一个判断，剔除无效数据
                templist.append(temp) # 提取第n列数据read_column(1)
        print("节次范围总数",len(templist))
        global number_weekrange
        number_weekrange=len(templist)
        return templist

#把类似于"A-B"的pair转为[A,B]的list，如果只是单个数字，则返回只有单个数字的list
def parse_range(pair_str):
    temp=pair_str.split("-")
    ret=[]
    for i in temp:
        ret.append(int(i))
    return ret

#返回一个课程项目的list套list：一个课程项目：[String course_section, [String week_section1,String week_section2 ……]]
#由于PDF是自动解析，没有divide一项
def parse_courselist_csv(csv_path):
    ret=[]
    tempA=get_column(csv_path,1)#获取课程范围的所有
    tempB=get_detail(csv_path)
    #保存一下，如果这个文件有问题，就直接写出xlsx替换掉pdf
    global list_courserange
    global list_weekrange
    list_courserange=tempA
    list_weekrange=tempB
    count=0
    for temp2 in tempA:
        count=count+1
        temp3=[]
        temp3.append(temp2)
        temp3.append(get_weekrange(tempB[count-1]))
        ret.append(temp3)
    #print(ret)#测试使用
    return ret

#返回一个课程项目的list套list：一个课程项目：[String course_section, [String week_section1,String week_section2 ……]]
#V3.1开始，最后一条记录为['divide',1,2,3,……]用于分隔每一周
def parse_courselist_xlsx(xlsx_path):
    ret=[]

    #xlsx操作
    coursebook=openpyxl.load_workbook(xlsx_path)
    coursebook_sheet=coursebook['Sheet1']

    #获取最大行数
    max_row=coursebook_sheet.max_row-1
    
    courserange=""
    weekrange=[]
    cell_id=""
    cell_id_before=""
    for i in range(1,max_row+1):
        cell_id="A"+str(i)
        cell_id_before="A"+str(i-1)
        if(coursebook_sheet[cell_id].value!=None):
            #添加上一个
            temp=[courserange,weekrange]
            ret.append(temp)
            weekrange=[]
            #更新courserange
            courserange=coursebook_sheet[cell_id].value
        cell_id="B"+str(i)
        #print("#",cell_id,get_weekrange(coursebook_sheet[cell_id].value))
        for tmp in get_weekrange(coursebook_sheet[cell_id].value):
            weekrange.append(tmp) 
         
    
    #添加最后一个
    temp=[courserange,weekrange]
    ret.append(temp)
    weekrange=[]
    #删除第一个空的
    ret.pop(0)
    #print("xlsx",ret)
    #添加最后一行，也就是divide。读取最多7个
    list_cellID=['A','B','C','D','E','F','G']
    list_divide=[]
    count=1
    for tmp in list_cellID:
        cell_id=list_cellID[count-1]+str(max_row+1)
        if(coursebook_sheet[cell_id].value!=None):
            list_divide.append(int(coursebook_sheet[cell_id].value))
        else:
            list_divide.append(0)
        
        count=count+1
    
    temp=["divide",list_divide]
    ret.append(temp)
    return ret

#输出一个人员,生成有课表
def output_member_full(sheet,department,name,courselist):
    pre=courselist[0]
    #调试一下，这里week总是碰到“8”
    week=1

    if_divide=courselist[len(courselist)-1][0]=="divide"
    if(if_divide):
        list_divide=courselist[len(courselist)-1][1]
        courselist.pop(len(courselist)-1)#删去divide记录
        accumulate=list_divide[0]
    currentrecord=1

    for temp in courselist:
        #拿到一条课程，先判断是否是一定有课，如果一定有课，则直接填
        #这里无需遍历课程周次范围，因为如果是一直有课，那么就有且只有一个周次段，也就是6-16或者6-17，只需要读取list第一个元素index=0

        tempA=parse_range(temp[0])#当前
        tempB=parse_range(pre[0])#前一个
         #判断是否切换到下一天，切换的同时写入全空的记录
        if(if_divide and accumulate<currentrecord):
            week=week+1
            accumulate=accumulate+list_divide[week-1]
            
        else:
            if(if_divide==False and tempA[1]<tempB[1]):#后面的比前面的小，那就是跳天了
                week=week+1

        
        if(temp[1][0]=="6-16" or temp[1][0]=="6-17"):
            temp_range=parse_range(temp[0])#解析出对应课程覆盖时间段
            for i in range(temp_range[0],temp_range[1]+1):
                set_record(sheet,department,name,week,i,"")
        else:
            temp_range=parse_range(temp[0])#解析出对应课程覆盖时间段
            for i in range(temp_range[0],temp_range[1]+1):
                set_record(sheet,department,name,week,i,temp[1])
        pre=temp#保存当前课程条目，供下一个比对确定是第几周
        currentrecord=currentrecord+1


#输出一个人员,生成无课表
def output_member_empty(sheet,department,name,courselist):
    pre=courselist[0]
    week=1
    mark=[0]*14
    if_divide=courselist[len(courselist)-1][0]=="divide"
    if(if_divide):
        list_divide=courselist[len(courselist)-1][1]
        courselist.pop(len(courselist)-1)#删去divide记录
        accumulate=list_divide[0]
    currentrecord=1
    #第一轮循环，先找出无课的，先写
    while(True):
        #拿到一条课程，先判断是否是一定有课，如果一定有课，则直接填
        #这里无需遍历课程周次范围，因为如果是一直有课，那么就有且只有一个周次段，也就是6-16或者6-17，只需要读取list第一个元素index=0
        if(currentrecord<=len(courselist)):
            temp=courselist[currentrecord-1]
        #这里是读取当前的(tempA)和前一个(tempB)的数据

        
        tempA=parse_range(temp[0])
        tempB=parse_range(pre[0])


        #判断是否切换到下一天，切换的同时写入全空的记录
        if(if_divide and accumulate<currentrecord):
            count=0
            for i in mark:
                count=count+1
                if(i==0): 
                    set_record(sheet,department,name,week,count,"")
            week=week+1
            mark=[0]*14#为下一天做准备
            if(week<=7):
                accumulate=accumulate+list_divide[week-1]
        else: 
            if(if_divide==False and tempA[1]<tempB[1]):#后面的比前面的小，那就是跳天了,跳天直接开始写当日的空
                #这里本来是0的，可是课表有改变，所以这个方法有风险
                count=0
                for i in mark:
                    count=count+1
                    if(i==0): 
                        set_record(sheet,department,name,week,count,"")
                week=week+1
                mark=[0]*14#为下一天做准备
        
        #判断什么时候跳出循环,多一次进行收尾
        if(currentrecord>len(courselist)):
            break
        
        
        temp_range=parse_range(temp[0])#解析出对应课程覆盖时间段
        #把有课的置为1
        for i in range(temp_range[0],temp_range[1]+1):
            mark[i-1]=1
        #顺便可以分析哪些是间断性的有课
        if(temp[1][0]!="6-16" and temp[1][0]!="6-17"):
            temp_range=parse_range(temp[0])#解析出对应课程覆盖时间段
            for i in range(temp_range[0],temp_range[1]+1):
                set_record(sheet,department,name,week,i,temp[1])
        pre=temp#保存当前课程条目，供下一个比对确定是第几周
        currentrecord=currentrecord+1

#数据库操作：写一条记录到对应Cell，week_range分两种，如果为空(请勿提交null)，则无，如果有范围，则记作：办公室 张三(5-7/8-9)
def set_record(sheet,department,name,week_num,section_num,week_range):
    #num均从1开始，方便观察;索引从0开始递增
    week_list=['C','D','E','F','G','H','I']
    section_list=list(range(2,16))#左开右闭一定注意
    msg_week_range=""
    if(week_range==""):
        msg_week_range=""
    else:
        msg_week_range="("
        count=0
        for temp in week_range:
            count=count+1
            if(count==len(week_range)):
                msg_week_range=msg_week_range+temp
            else:
                msg_week_range=msg_week_range+temp+"/"
            
        msg_week_range=msg_week_range+")"
    
    cell_id=week_list[week_num-1]+str(section_list[section_num-1])
    if sheet[cell_id].value==None:
        sheet[cell_id]=department+' '+name+msg_week_range+'，'
    else:
        sheet[cell_id]=sheet[cell_id].value+department+' '+name+msg_week_range+'，'

#准备工作
def welcome():
    print('-----------------------------')
    print('无课表自动生成系统V2.0，西南大学人工智能学院团委办公室部门制作，作者Blue16，个人主页blue16.cn')
    print('由于数据有误，我在10月8日当天完成了核心代码重写，优化了算法逻辑,同时让程序能支持更多的功能，比如生成有课表（狗头）')
    input('请将所有数据按照格式要求放入InputTable文件夹中，程序解析结束可以指定输出位置。按回车继续')
    print('开始处理')
    print('-----------------------------')

#-------测试代码开始
#wb=openpyxl.load_workbook('output.xlsx')
#wb_sheet=wb['Sheet']
#path=os.getcwd()#D:\桌面\无课表
#output_member2(wb_sheet,"办公室","孙文博",parse_courselist("D:\\Programming\\freetimetable-generate\\InputTable\\办公室\\孙文博.csv"))
#wb.save("mytest.xlsx")#注意保存
#-------测试代码结束

#搜索对应部门成员然后直接写入
def search_department_member(path,department):
    member_file_list=os.listdir(path+'\\InputTable\\'+department)
    member_list=[]
    type="pdf"
    for temp in member_file_list:
        temp2=temp.split('.')
        if(temp2[1]=='pdf' or temp2[1]=='xlsx'):
            member_list.append(temp)

    print('查找到的成员文件:',member_list)
    for temp in member_list:
        temp2=temp.split('(')
        print('当前写入部门：',department,"当前写入人员：",temp2[0])
        #准备工作
        sourcepath=path+'\\InputTable\\'+department+'\\'+temp
        destpath=path+'\\InputTable\\'+department+'\\'+temp2[0]+'.csv'
        temp3=temp.split('.')
        type=temp3[1]
        #如果想体验生成有课表，请调用output_member_full，参数完全相同
        if(type=="pdf"):
            convert_to_csv(sourcepath,destpath)
            output_member_empty(wb_sheet,department,temp2[0],parse_courselist_csv(destpath))
        if(type=="xlsx"):
            output_member_empty(wb_sheet,department,temp2[0],parse_courselist_xlsx(sourcepath))
        
        #print("@",number_detail,number_weekrange)
        if(number_detail!=number_weekrange):
            list_failed.append(department+" "+temp2[0])
            os.remove(sourcepath)
            wb1=openpyxl.Workbook()
            wb_sheet1=wb1.create_sheet("Sheet1")
            count=0
            for temp in list_courserange:
                count=count+1
                cell_id="A"+str(count)
                wb_sheet1[cell_id]=temp
            count=0
            for temp in list_weekrange:
                count=count+1
                cell_id="B"+str(count)
                wb_sheet1[cell_id]=temp
            #wb1.remove_sheet("Sheet")
            wb1.save(sourcepath.replace(".pdf",".xlsx"))

#正式代码
welcome()
wb=openpyxl.load_workbook('output.xlsx')
wb_sheet=wb['Sheet']
path=os.getcwd()#D:\桌面\无课表
department_list=os.listdir(path+'\\InputTable')
print('查找到的部门：',department_list)
for temp in department_list:
    search_department_member(path,temp)
if(len(list_failed)!=0):
    print("@@@警告：下列成员需要调整对应xlsx文件，文件已经替换成xlsx，请调整后再次执行程序",list_failed)
    exit()
outputpath=input('解析完成，请输入要保存的文件名,然后回车：（可以是路径）')
while outputpath=='':
    outputpath=input('文件名无效，请输入文件名后回车：（可以是路径）')
wb.save(outputpath)#注意保存
print('文件已经保存到'+outputpath+'。完成，撒花')

